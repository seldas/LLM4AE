#!/usr/bin/env python3
"""Persist authoritative FAERS case-series assignments in ``dataset.db``.

Uses the four original FDA query case series workbooks in
``publication/Datasets/Original_series`` as the canonical ground truth:
  - 6_Final_QQ_Azacitidine_QT prolongation.xlsx      -> Azacitidine-QT
  - 8_Final_QQ_Baricitinib_Hypersensitivity.xlsx     -> Baricitinib-Hypersensitivity
  - 4_Final_QQ_Erenumab_Stroke.xlsx                  -> Erenumab-Stroke
  - 2_Final_QQ_Tramadol_Hypoglycemia.xlsx            -> Tramadol-Hypoglycemia

All 829 FAERS reports in ``dataset.db`` are matched by FAERS Case # and
Version Number.
"""

from __future__ import annotations

import argparse
import csv
import os
import sqlite3
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


CASE_SERIES_SPECS = [
    {
        "series": "Azacitidine-QT",
        "file": "6_Final_QQ_Azacitidine_QT prolongation.xlsx",
    },
    {
        "series": "Baricitinib-Hypersensitivity",
        "file": "8_Final_QQ_Baricitinib_Hypersensitivity.xlsx",
    },
    {
        "series": "Erenumab-Stroke",
        "file": "4_Final_QQ_Erenumab_Stroke.xlsx",
    },
    {
        "series": "Tramadol-Hypoglycemia",
        "file": "2_Final_QQ_Tramadol_Hypoglycemia.xlsx",
    },
]

CASE_SERIES = tuple(spec["series"] for spec in CASE_SERIES_SPECS)


def parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--database",
        type=Path,
        default=repo_root / "publication" / "dataset.db",
        help="Path to SQLite database (default: publication/dataset.db)",
    )
    parser.add_argument(
        "--series-dir",
        type=Path,
        default=repo_root / "publication" / "Datasets" / "Original_series",
        help="Path to directory containing original Excel case series",
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=(
            repo_root
            / "publication"
            / "results"
            / "faers_case_series_classification.csv"
        ),
        help="Path to output audit CSV",
    )
    parser.add_argument(
        "--exclude-empty-narrative",
        action="store_true",
        help="Set include_in_loo=0 for doc 3047591-1 (pre-Nov 1997 placeholder text)",
    )
    return parser.parse_args()


def load_original_series(
    series_dir: Path,
) -> Tuple[Dict[str, dict], Dict[str, List[str]]]:
    """Extract case mappings and metadata from the four Excel workbooks."""
    mapping: Dict[str, dict] = {}
    duplicates: Dict[str, List[str]] = {}

    for spec in CASE_SERIES_SPECS:
        series_name = spec["series"]
        file_path = series_dir / spec["file"]
        if not file_path.is_file():
            raise FileNotFoundError(f"Original series file not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        if "Line Listing of Cases" not in wb.sheetnames:
            raise ValueError(f"Sheet 'Line Listing of Cases' not in {file_path}")

        sheet = wb["Line Listing of Cases"]
        # Header is row 3: col 2 is Case #, col 3 is Version Number
        for r in range(4, sheet.max_row + 1):
            case_no = sheet.cell(row=r, column=2).value
            ver_no = sheet.cell(row=r, column=3).value
            if case_no is None:
                continue
            case_str = str(case_no).strip()
            ver_str = str(ver_no).strip() if ver_no is not None else ""
            doc_id = f"{case_str}-{ver_str}"

            if doc_id in mapping:
                if doc_id not in duplicates:
                    duplicates[doc_id] = [mapping[doc_id]["case_series"]]
                duplicates[doc_id].append(series_name)
                continue

            ps_drug = sheet.cell(row=r, column=7).value
            pts = sheet.cell(row=r, column=10).value

            mapping[doc_id] = {
                "doc_id": doc_id,
                "case_series": series_name,
                "excel_file": spec["file"],
                "excel_row": r,
                "case_no": case_str,
                "version_no": ver_str,
                "primary_suspect": str(ps_drug).strip() if ps_drug else "",
                "pts": str(pts).strip() if pts else "",
            }

    if duplicates:
        raise ValueError(f"Duplicate doc_ids across series workbooks: {duplicates}")

    return mapping, duplicates


def main() -> None:
    args = parse_args()
    database_path = args.database.resolve()
    series_dir = args.series_dir.resolve()
    output_csv = args.output_csv.resolve()

    if not database_path.is_file():
        raise FileNotFoundError(f"Database not found: {database_path}")

    excel_mapping, _ = load_original_series(series_dir)

    with sqlite3.connect(database_path) as connection:
        connection.execute("PRAGMA foreign_keys = ON")
        faers_docs = [
            str(row[0])
            for row in connection.execute(
                "SELECT doc_id FROM documents WHERE dataset = 'FAERS' ORDER BY doc_id"
            )
        ]
        faers_set = set(faers_docs)

        # Check for missing docs
        missing_in_excel = faers_set - set(excel_mapping)
        if missing_in_excel:
            raise ValueError(
                f"FAERS documents not found in Excel files ({len(missing_in_excel)}): "
                f"{sorted(missing_in_excel)[:10]}"
            )

        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS faers_case_series (
                doc_id             TEXT PRIMARY KEY REFERENCES documents(doc_id),
                case_series        TEXT,
                include_in_loo     INTEGER NOT NULL
                                   CHECK (include_in_loo IN (0, 1)),
                assignment_source  TEXT NOT NULL,
                assignment_note    TEXT,
                CHECK (case_series IS NULL OR case_series IN (
                    'Azacitidine-QT',
                    'Tramadol-Hypoglycemia',
                    'Baricitinib-Hypersensitivity',
                    'Erenumab-Stroke'
                )),
                CHECK (
                    (include_in_loo = 1 AND case_series IS NOT NULL) OR
                    include_in_loo = 0
                )
            )
            """
        )
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_faers_case_series_name "
            "ON faers_case_series(case_series)"
        )
        connection.execute("DELETE FROM faers_case_series")

        db_rows = []
        csv_rows = []
        distribution: Counter[str] = Counter()
        loo_distribution: Counter[str] = Counter()

        for doc_id in faers_docs:
            info = excel_mapping[doc_id]
            series = info["case_series"]
            distribution[series] += 1

            if doc_id == "3047591-1" and args.exclude_empty_narrative:
                include_in_loo = 0
                note = (
                    f"From {info['excel_file']} (row {info['excel_row']}); "
                    f"PS: {info['primary_suspect']}. Excluded from LOO: pre-Nov 1997 placeholder text."
                )
            else:
                include_in_loo = 1
                note = f"From {info['excel_file']} (row {info['excel_row']}); PS: {info['primary_suspect']}"

            if include_in_loo == 1:
                loo_distribution[series] += 1

            source = "fda_original_series_excel"
            db_rows.append((doc_id, series, include_in_loo, source, note))

            csv_rows.append([
                doc_id,
                series,
                "high",
                "No" if include_in_loo == 1 else "Yes",
                note,
                1000,
                "",
                0,
                1000,
                info["primary_suspect"],
                "",
                "",
                info["pts"],
                "",
                f"{series}=1000",
            ])

        connection.executemany(
            """
            INSERT INTO faers_case_series (
                doc_id, case_series, include_in_loo,
                assignment_source, assignment_note
            ) VALUES (?, ?, ?, ?, ?)
            """,
            db_rows,
        )

    # Write CSV audit
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    header = [
        "doc_id",
        "assigned_series",
        "confidence",
        "review_required",
        "review_reason",
        "top_score",
        "runner_up_series",
        "runner_up_score",
        "score_margin",
        "suspect_drug_evidence",
        "other_drug_evidence",
        "text_drug_evidence",
        "annotated_event_evidence",
        "text_event_evidence",
        "all_series_scores",
    ]
    with output_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(header)
        writer.writerows(csv_rows)

    print(f"Authoritative FAERS case-series persisted in: {database_path}")
    print(f"Total FAERS records: {len(faers_docs):,}")
    print("Case series distribution (All):")
    for s in CASE_SERIES:
        print(f"  {s}: {distribution[s]:,}")
    print("LOO cohort distribution (include_in_loo=1):")
    for s in CASE_SERIES:
        print(f"  {s}: {loo_distribution[s]:,}")
    if args.exclude_empty_narrative:
        print("  Excluded from LOO: 1 (3047591-1)")
    print(f"Audit CSV written to: {output_csv}")


if __name__ == "__main__":
    main()

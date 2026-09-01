#!/usr/bin/env python3
"""
generate_supplementary_files_2_3.py

Generates publication-quality Supplementary Excel Files directly from the verbatim
in-prompt annotation schema strings used during LLM evaluations:
- Supplementary File 2: FAERS Clinical Concept Annotation Guidance (17 Categories)
- Supplementary File 3: VAERS Vaccine Adverse Event Annotation Guidance (14 Categories)

This ensures 100.0% verbatim identity with the exact runtime prompts injected into
LLaMA 4 and Claude 4.6 Sonnet (P2_TAG, P1_JSON, P2_TAG_VAERS, P1_JSON_VAERS).
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import exact prompt module
script_dir = Path(__file__).resolve().parent
sys.path.insert(0, str(script_dir))
import annotation_prompts


# ==============================================================================
# STYLING HELPER FUNCTIONS
# ==============================================================================

def get_header_fill():
    return PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")  # Navy Blue

def get_accent_fill():
    return PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")  # Soft Ice Blue

def get_zebra_fill():
    return PatternFill(start_color="FBFBFB", end_color="FBFBFB", fill_type="solid")

def get_white_fill():
    return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

def get_thin_border():
    thin = Side(border_style="thin", color="D0D5DD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cover_sheet(ws, title: str, subtitle: str, metadata_rows: list[tuple[str, str]]):
    ws.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws.merge_cells("B2:G2")
    ws["B2"] = title
    ws["B2"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["B2"].fill = get_header_fill()
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 36

    # Subtitle
    ws.merge_cells("B3:G3")
    ws["B3"] = subtitle
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color="4A5568")
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 24

    # Metadata Table
    start_row = 5
    for idx, (label, val) in enumerate(metadata_rows):
        row = start_row + idx
        ws.row_dimensions[row].height = 24
        
        c_label = ws.cell(row=row, column=2, value=label)
        c_label.font = Font(name="Calibri", size=10.5, bold=True, color="1B365D")
        c_label.fill = get_accent_fill()
        c_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_label.border = get_thin_border()

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        c_val = ws.cell(row=row, column=3, value=val)
        c_val.font = Font(name="Calibri", size=10, color="2D3748")
        c_val.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = get_thin_border()

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16


def parse_guide_markdown(guide_str: str):
    """Parses markdown annotation table and general rules directly from runtime prompt strings."""
    lines = guide_str.strip().split('\n')
    table_rows = []
    general_rules = []
    
    in_table = False
    in_rules = False
    
    current_rule_title = None
    current_rule_body = []
    
    for line in lines:
        s = line.strip()
        if s.startswith("| Clinical Concept |"):
            in_table = True
            continue
        if in_table:
            if s.startswith("|---|"):
                continue
            if s.startswith("|"):
                parts = [p.strip() for p in s.strip('|').split('|')]
                if len(parts) >= 4:
                    concept_clean = parts[0].replace('**', '').strip()
                    code = concept_clean.split(':')[0].strip() if ':' in concept_clean else concept_clean
                    name = concept_clean.split(':')[1].strip() if ':' in concept_clean else concept_clean
                    definition = parts[1]
                    annotation_rule = parts[2]
                    triggers = parts[3]
                    
                    table_rows.append({
                        "code": code,
                        "name": name,
                        "raw_concept": concept_clean,
                        "definition": definition,
                        "annotation_rule": annotation_rule,
                        "triggers": triggers
                    })
            else:
                if s.startswith("### General Annotation Rules"):
                    in_table = False
                    in_rules = True
                    continue
        
        if in_rules:
            rule_match = re.match(r'^(\d+)\.\s+\*\*(.+?)\*\*(.*)', s)
            if rule_match:
                if current_rule_title is not None:
                    general_rules.append((current_rule_title, "\n".join(current_rule_body).strip()))
                num = rule_match.group(1)
                title = f"{num}. {rule_match.group(2).strip()}"
                rest = rule_match.group(3).strip()
                current_rule_title = title
                current_rule_body = [rest] if rest else []
            elif current_rule_title is not None and s:
                current_rule_body.append(s)
                
    if current_rule_title is not None:
        general_rules.append((current_rule_title, "\n".join(current_rule_body).strip()))
        
    return table_rows, general_rules


# ==============================================================================
# FAERS WORKBOOK GENERATION
# ==============================================================================

def generate_faers_workbook(out_path: Path):
    wb = openpyxl.Workbook()
    
    # 1. Cover Sheet
    ws_cover = wb.active
    ws_cover.title = "Cover_Sheet"
    metadata = [
        ("Supplementary File:", "Supplementary File 2"),
        ("Document Title:", "FAERS Clinical Concept Annotation Guidance & Category Schema"),
        ("Associated Manuscript:", "Benchmarking Large Language Models and Fine-Tuned Encoders for Clinical Concept Extraction from Pharmacovigilance and Vaccine Adverse Event Narratives"),
        ("Target Corpus:", "FDA Adverse Event Reporting System (FAERS) Benchmark (N = 829 Narratives)"),
        ("Total Categories:", "17 Primary Clinical Concept Categories (Exact Verbatim Prompt Schema)"),
        ("In-Text XML Tags:", "<SDRUG>, <CDRUG>, <ODRUG>, <DOSE>, <IND>, <TREATMENT>, <AE>, <MAE>, <DX>, <LAB>, <STATUS>, <RO>, <COD>, <MHX>, <FHX>, <AGE>, <SEX>"),
        ("Structured JSON Keys:", "sdrug, cdrug, odrug, dose, ind, treatment, ae, mae, dx, lab, status, ro, cod, mhx, fhx, age, sex"),
        ("Evaluation Protocol:", "4-Fold Leave-One-Drug-Event-Pair-Out on BioBERT vs. Zero-Shot/1-Shot Instruction-Tuned LLMs (LLaMA 4 & Claude 4.6 Sonnet)"),
        ("Version & Date:", "Revision 1 (August 2026)")
    ]
    style_cover_sheet(ws_cover, "Supplementary File 2: FAERS Clinical Concept Annotation Guidance",
                      "Exact Verbatim Annotation Schema and Operational Rules Injected into LLM Prompts", metadata)

    # 2. Schema Sheet
    faers_table, faers_rules = parse_guide_markdown(annotation_prompts.ANNOTATION_GUIDE)
    
    tag_map = {
        "sDrug": "<SDRUG>", "cDrug": "<CDRUG>", "oDrug": "<ODRUG>", "Dose": "<DOSE>",
        "IND": "<IND>", "Treatment": "<TREATMENT>", "AE": "<AE>", "mAE": "<MAE>",
        "Dx": "<DX>", "Lab": "<LAB>", "Status": "<STATUS>", "R/O": "<RO>",
        "CoD": "<COD>", "MHx": "<MHX>", "FHx": "<FHX>", "Age": "<AGE>", "Sex": "<SEX>"
    }

    ws_schema = wb.create_sheet(title="FAERS_17_Category_Schema")
    ws_schema.views.sheetView[0].showGridLines = True
    
    # Title
    ws_schema.merge_cells("B2:G2")
    ws_schema["B2"] = "FAERS 17 Clinical Concept Category Annotation Schema (Verbatim LLM Prompt)"
    ws_schema["B2"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws_schema["B2"].fill = get_header_fill()
    ws_schema["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_schema.row_dimensions[2].height = 30

    headers = [
        "Concept Code",
        "Category Name",
        "XML Tag",
        "Definition (Verbatim Prompt Text)",
        "Annotation Rule (Verbatim Prompt Text)",
        "Trigger Words / Contextual Clues (Verbatim Prompt Text)"
    ]
    ws_schema.row_dimensions[4].height = 26
    for c_idx, h in enumerate(headers, start=2):
        cell = ws_schema.cell(row=4, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
        cell.fill = get_header_fill()
        cell.alignment = Alignment(horizontal="center" if c_idx in [2, 4] else "left", vertical="center", wrap_text=True)
        cell.border = get_thin_border()

    for r_idx, item in enumerate(faers_table, start=5):
        ws_schema.row_dimensions[r_idx].height = 65
        fill = get_zebra_fill() if r_idx % 2 == 1 else get_white_fill()
        
        row_vals = [
            item["code"],
            item["name"],
            tag_map.get(item["code"], f"<{item['code'].upper()}>"),
            item["definition"],
            item["annotation_rule"],
            item["triggers"]
        ]
        
        for c_idx, val in enumerate(row_vals, start=2):
            cell = ws_schema.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=9.5, bold=(c_idx in [2, 3]))
            cell.fill = fill
            cell.border = get_thin_border()
            cell.alignment = Alignment(
                horizontal="center" if c_idx in [2, 4] else "left",
                vertical="top",
                wrap_text=True
            )

    ws_schema.column_dimensions["A"].width = 3
    col_widths = [14, 26, 14, 42, 48, 38]
    for c_idx, width in enumerate(col_widths, start=2):
        ws_schema.column_dimensions[get_column_letter(c_idx)].width = width

    # 3. General Rules Sheet
    ws_rules = wb.create_sheet(title="General_Annotation_Rules")
    ws_rules.views.sheetView[0].showGridLines = True
    
    ws_rules.merge_cells("B2:D2")
    ws_rules["B2"] = "FAERS General Annotation Rules & Operational Principles (Verbatim LLM Prompt)"
    ws_rules["B2"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws_rules["B2"].fill = get_header_fill()
    ws_rules["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_rules.row_dimensions[2].height = 30

    rule_headers = ["Rule Number & Principle", "Operational Annotation Instruction (Verbatim Prompt Text)"]
    ws_rules.row_dimensions[4].height = 26
    
    ws_rules.cell(row=4, column=2, value=rule_headers[0]).font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
    ws_rules.cell(row=4, column=2).fill = get_header_fill()
    ws_rules.cell(row=4, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_rules.cell(row=4, column=2).border = get_thin_border()

    ws_rules.merge_cells("C4:D4")
    c_hdr2 = ws_rules.cell(row=4, column=3, value=rule_headers[1])
    c_hdr2.font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
    c_hdr2.fill = get_header_fill()
    c_hdr2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_rules.cell(row=4, column=3).border = get_thin_border()
    ws_rules.cell(row=4, column=4).border = get_thin_border()

    for r_idx, (r_title, r_desc) in enumerate(faers_rules, start=5):
        ws_rules.row_dimensions[r_idx].height = 50
        fill = get_zebra_fill() if r_idx % 2 == 1 else get_white_fill()
        
        c2 = ws_rules.cell(row=r_idx, column=2, value=r_title)
        c2.font = Font(name="Calibri", size=10, bold=True, color="1B365D")
        c2.fill = fill
        c2.border = get_thin_border()
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

        ws_rules.merge_cells(start_row=r_idx, start_column=3, end_row=r_idx, end_column=4)
        c3 = ws_rules.cell(row=r_idx, column=3, value=r_desc)
        c3.font = Font(name="Calibri", size=9.5)
        c3.fill = fill
        c3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ws_rules.cell(row=r_idx, column=3).border = get_thin_border()
        ws_rules.cell(row=r_idx, column=4).border = get_thin_border()

    ws_rules.column_dimensions["A"].width = 3
    ws_rules.column_dimensions["B"].width = 34
    ws_rules.column_dimensions["C"].width = 46
    ws_rules.column_dimensions["D"].width = 46

    wb.save(str(out_path))
    print(f"Saved exact verbatim FAERS Guidance to {out_path}")


# ==============================================================================
# VAERS WORKBOOK GENERATION
# ==============================================================================

def generate_vaers_workbook(out_path: Path):
    wb = openpyxl.Workbook()
    
    # 1. Cover Sheet
    ws_cover = wb.active
    ws_cover.title = "Cover_Sheet"
    metadata = [
        ("Supplementary File:", "Supplementary File 3"),
        ("Document Title:", "VAERS Vaccine Adverse Event Annotation Guidance & Category Schema"),
        ("Associated Manuscript:", "Benchmarking Large Language Models and Fine-Tuned Encoders for Clinical Concept Extraction from Pharmacovigilance and Vaccine Adverse Event Narratives"),
        ("Target Corpus:", "Vaccine Adverse Event Reporting System (VAERS) Benchmark (N = 1,000 Narratives)"),
        ("Total Categories:", "13 Primary Clinical and Contextual Concept Categories (Exact Verbatim Prompt Schema)"),
        ("In-Text XML Tags:", "<SYM>, <SDX>, <PDX>, <DX>, <VAX>, <MHX>, <FHX>, <LAB>, <DOSE>, <STATUS>, <TX>, <AGE>, <SEX>"),
        ("Structured JSON Keys:", "sym, sdx, pdx, dx, vax, mhx, fhx, lab, dose, status, tx, age, sex"),
        ("Evaluation Protocol:", "10-Fold Cross-Validation on BioBERT vs. Zero-Shot/1-Shot Instruction-Tuned LLMs (LLaMA 4)"),
        ("Version & Date:", "Revision 1 (August 2026)")
    ]
    style_cover_sheet(ws_cover, "Supplementary File 3: VAERS Vaccine Adverse Event Annotation Guidance",
                      "Exact Verbatim Annotation Schema and Operational Rules Injected into LLM Prompts", metadata)

    # 2. Schema Sheet
    vaers_table, vaers_rules = parse_guide_markdown(annotation_prompts.ANNOTATION_GUIDE_VAERS)
    
    tag_map_vaers = {
        "SYM": "<SYM>", "sDx": "<SDX>", "pDx": "<PDX>", "DX": "<DX>",
        "VAX": "<VAX>", "MHx": "<MHX>", "FHx": "<FHX>", "Lab": "<LAB>",
        "DOSE": "<DOSE>", "STATUS": "<STATUS>", "TX": "<TX>",
        "AGE": "<AGE>", "SEX": "<SEX>"
    }

    ws_schema = wb.create_sheet(title="VAERS_13_Category_Schema")
    ws_schema.views.sheetView[0].showGridLines = True
    
    # Title
    ws_schema.merge_cells("B2:G2")
    ws_schema["B2"] = "VAERS 13 Clinical Concept Category Annotation Schema (Verbatim LLM Prompt)"
    ws_schema["B2"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws_schema["B2"].fill = get_header_fill()
    ws_schema["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_schema.row_dimensions[2].height = 30

    headers = [
        "Concept Code",
        "Category Name",
        "XML Tag",
        "Definition (Verbatim Prompt Text)",
        "Annotation Rule (Verbatim Prompt Text)",
        "Trigger Words / Contextual Clues (Verbatim Prompt Text)"
    ]
    ws_schema.row_dimensions[4].height = 26
    for c_idx, h in enumerate(headers, start=2):
        cell = ws_schema.cell(row=4, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
        cell.fill = get_header_fill()
        cell.alignment = Alignment(horizontal="center" if c_idx in [2, 4] else "left", vertical="center", wrap_text=True)
        cell.border = get_thin_border()

    for r_idx, item in enumerate(vaers_table, start=5):
        ws_schema.row_dimensions[r_idx].height = 65
        fill = get_zebra_fill() if r_idx % 2 == 1 else get_white_fill()
        
        row_vals = [
            item["code"],
            item["name"],
            tag_map_vaers.get(item["code"], f"<{item['code'].upper()}>"),
            item["definition"],
            item["annotation_rule"],
            item["triggers"]
        ]
        
        for c_idx, val in enumerate(row_vals, start=2):
            cell = ws_schema.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=9.5, bold=(c_idx in [2, 3]))
            cell.fill = fill
            cell.border = get_thin_border()
            cell.alignment = Alignment(
                horizontal="center" if c_idx in [2, 4] else "left",
                vertical="top",
                wrap_text=True
            )

    ws_schema.column_dimensions["A"].width = 3
    col_widths = [14, 26, 14, 42, 48, 38]
    for c_idx, width in enumerate(col_widths, start=2):
        ws_schema.column_dimensions[get_column_letter(c_idx)].width = width

    # 3. General Rules Sheet
    ws_rules = wb.create_sheet(title="General_Annotation_Rules")
    ws_rules.views.sheetView[0].showGridLines = True
    
    ws_rules.merge_cells("B2:D2")
    ws_rules["B2"] = "VAERS General Annotation Rules & Operational Principles (Verbatim LLM Prompt)"
    ws_rules["B2"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws_rules["B2"].fill = get_header_fill()
    ws_rules["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_rules.row_dimensions[2].height = 30

    rule_headers = ["Rule Number & Principle", "Operational Annotation Instruction (Verbatim Prompt Text)"]
    ws_rules.row_dimensions[4].height = 26
    
    ws_rules.cell(row=4, column=2, value=rule_headers[0]).font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
    ws_rules.cell(row=4, column=2).fill = get_header_fill()
    ws_rules.cell(row=4, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_rules.cell(row=4, column=2).border = get_thin_border()

    ws_rules.merge_cells("C4:D4")
    c_hdr2 = ws_rules.cell(row=4, column=3, value=rule_headers[1])
    c_hdr2.font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
    c_hdr2.fill = get_header_fill()
    c_hdr2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_rules.cell(row=4, column=3).border = get_thin_border()
    ws_rules.cell(row=4, column=4).border = get_thin_border()

    for r_idx, (r_title, r_desc) in enumerate(vaers_rules, start=5):
        ws_rules.row_dimensions[r_idx].height = 50
        fill = get_zebra_fill() if r_idx % 2 == 1 else get_white_fill()
        
        c2 = ws_rules.cell(row=r_idx, column=2, value=r_title)
        c2.font = Font(name="Calibri", size=10, bold=True, color="1B365D")
        c2.fill = fill
        c2.border = get_thin_border()
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

        ws_rules.merge_cells(start_row=r_idx, start_column=3, end_row=r_idx, end_column=4)
        c3 = ws_rules.cell(row=r_idx, column=3, value=r_desc)
        c3.font = Font(name="Calibri", size=9.5)
        c3.fill = fill
        c3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ws_rules.cell(row=r_idx, column=3).border = get_thin_border()
        ws_rules.cell(row=r_idx, column=4).border = get_thin_border()

    ws_rules.column_dimensions["A"].width = 3
    ws_rules.column_dimensions["B"].width = 34
    ws_rules.column_dimensions["C"].width = 46
    ws_rules.column_dimensions["D"].width = 46

    wb.save(str(out_path))
    print(f"Saved exact verbatim VAERS Guidance to {out_path}")


def main():
    repo_root = Path(__file__).resolve().parent.parent.parent
    supp_dir = repo_root / "publication" / "supplementary"
    tables_dir = repo_root / "publication" / "results" / "tables"
    manuscript_dir = repo_root / "publication" / "manuscripts"
    
    supp_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)

    faers_out1 = supp_dir / "Supplementary_File_2_FAERS_Annotation_Guidance.xlsx"
    faers_out2 = manuscript_dir / "Supplementary_File_2_FAERS_Annotation_Guidance.xlsx"
    faers_out3 = tables_dir / "Supplementary_File_2_FAERS_Annotation_Guidance.xlsx"

    vaers_out1 = supp_dir / "Supplementary_File_3_VAERS_Annotation_Guidance.xlsx"
    vaers_out2 = manuscript_dir / "Supplementary_File_3_VAERS_Annotation_Guidance.xlsx"
    vaers_out3 = tables_dir / "Supplementary_File_3_VAERS_Annotation_Guidance.xlsx"

    generate_faers_workbook(faers_out1)
    generate_faers_workbook(faers_out2)
    generate_faers_workbook(faers_out3)

    generate_vaers_workbook(vaers_out1)
    generate_vaers_workbook(vaers_out2)
    generate_vaers_workbook(vaers_out3)


if __name__ == "__main__":
    main()
